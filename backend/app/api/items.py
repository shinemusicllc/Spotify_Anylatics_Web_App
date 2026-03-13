"""Items endpoints — query stored items."""

import asyncio
from collections import defaultdict
from datetime import datetime
import io
import uuid

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.crawl_job import CrawlJob
from app.models.item import Item
from app.models.metrics_snapshot import MetricsSnapshot
from app.models.raw_response import RawResponse
from app.models.user import User
from app.schemas.item import ItemListResponse, ItemResponse
from app.services.auth import get_current_user
from app.services import spotify_client

router = APIRouter()


EXPORT_ACTIONS = {
    "listview-excel",
    "playlist-type3",
    "album-type0",
    "track-offline",
}
EXPORT_FORMATS = {"json", "txt", "xlsx"}


class ItemExportRequest(BaseModel):
    action: str
    format: str = "json"
    item_ids: list[str] = Field(default_factory=list)
    deep_fetch: bool = False


def _spotify_url(item_type: str, spotify_id: str) -> str:
    return f"https://open.spotify.com/{item_type}/{spotify_id}"


def _extract_owner_url(item: Item, raw_data: dict | None) -> str | None:
    if isinstance(raw_data, dict):
        owner_url = raw_data.get("owner_url")
        if isinstance(owner_url, str) and owner_url.strip():
            return owner_url.strip()

        if item.item_type == "track":
            artists = raw_data.get("artists") or []
            if isinstance(artists, list) and artists:
                first_artist = artists[0] or {}
                artist_id = first_artist.get("spotify_id") or first_artist.get("id")
                if artist_id:
                    return _spotify_url("artist", artist_id)

        if item.item_type == "album":
            artists = raw_data.get("artists") or []
            if isinstance(artists, list) and artists:
                first_artist = artists[0] or {}
                artist_id = first_artist.get("spotify_id") or first_artist.get("id")
                if artist_id:
                    return _spotify_url("artist", artist_id)

            tracks = raw_data.get("tracks") or []
            if isinstance(tracks, list) and tracks:
                first_track = tracks[0] or {}
                track_artists = first_track.get("artists") or []
                if isinstance(track_artists, list) and track_artists:
                    first_artist = track_artists[0] or {}
                    artist_id = first_artist.get("spotify_id") or first_artist.get("id")
                    if artist_id:
                        return _spotify_url("artist", artist_id)

    if item.item_type == "artist":
        return _spotify_url("artist", item.spotify_id)
    return None


def _extract_artist_names(raw_data: dict | None) -> list[str]:
    if not isinstance(raw_data, dict):
        return []

    artist_names = raw_data.get("artist_names")
    if isinstance(artist_names, list):
        return [
            name.strip()
            for name in artist_names
            if isinstance(name, str) and name.strip()
        ]

    artists = raw_data.get("artists")
    if isinstance(artists, list):
        names: list[str] = []
        for artist in artists:
            if not isinstance(artist, dict):
                continue
            name = artist.get("name")
            if isinstance(name, str) and name.strip():
                names.append(name.strip())
        return names

    return []


def _build_album_export_tracks(raw_data: dict | None) -> list[dict]:
    if not isinstance(raw_data, dict):
        return []

    tracks = raw_data.get("tracks")
    if not isinstance(tracks, list):
        return []

    export_rows: list[dict] = []
    for track in tracks:
        if not isinstance(track, dict):
            continue

        track_name = track.get("name")
        if not isinstance(track_name, str) or not track_name.strip():
            continue

        artist_names = _extract_artist_names(track)
        spotify_url = track.get("spotify_url")
        if not isinstance(spotify_url, str) or not spotify_url.strip():
            spotify_id = track.get("spotify_id")
            spotify_url = (
                _spotify_url("track", spotify_id.strip())
                if isinstance(spotify_id, str) and spotify_id.strip()
                else None
            )

        export_rows.append(
            {
                "artist_names": ", ".join(artist_names) if artist_names else "-",
                "track_name": track_name.strip(),
                "spotify_url": spotify_url,
            }
        )

    return export_rows


def _normalize_group_name(value: str | None) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    if "::" in text:
        maybe_owner, possible_name = text.split("::", 1)
        try:
            uuid.UUID(maybe_owner.strip())
            text = possible_name.strip()
        except ValueError:
            pass
    return text.lower()


def _format_added_date(value) -> str | None:
    if value is None:
        return None
    return value.strftime("%d/%m %H:%M")


def _compute_delta(current: int | None, previous: int | None) -> int | None:
    if current is None or previous is None:
        return None
    return current - previous


async def _load_latest_raw_data(db: AsyncSession, items: list[Item]) -> dict[str, dict]:
    spotify_ids = [item.spotify_id for item in items if item.spotify_id]
    if not spotify_ids:
        return {}

    result = await db.execute(
        select(RawResponse)
        .where(RawResponse.spotify_id.in_(spotify_ids))
        .order_by(RawResponse.captured_at.desc())
    )
    rows = result.scalars().all()
    raw_map: dict[str, dict] = {}
    for row in rows:
        if row.spotify_id not in raw_map and isinstance(row.response_data, dict):
            raw_map[row.spotify_id] = row.response_data
    return raw_map


async def _load_recent_snapshots(db: AsyncSession, items: list[Item]) -> dict:
    item_ids = [item.id for item in items if item.id]
    if not item_ids:
        return {}

    result = await db.execute(
        select(MetricsSnapshot)
        .where(MetricsSnapshot.item_id.in_(item_ids))
        .order_by(MetricsSnapshot.item_id, MetricsSnapshot.captured_at.desc())
    )
    rows = result.scalars().all()

    snapshot_map: dict = defaultdict(list)
    for row in rows:
        bucket = snapshot_map[row.item_id]
        if len(bucket) < 2:
            bucket.append(row)
    return snapshot_map


async def _load_item_users(db: AsyncSession, items: list[Item]) -> dict[str, User]:
    user_ids = [item.user_id for item in items if item.user_id]
    if not user_ids:
        return {}

    result = await db.execute(select(User).where(User.id.in_(user_ids)))
    users = result.scalars().all()
    return {str(user.id): user for user in users}


async def _delete_raw_if_unreferenced(
    db: AsyncSession,
    spotify_ids: list[str],
    excluded_item_ids: list[uuid.UUID] | None = None,
) -> None:
    unique_ids = [sid for sid in dict.fromkeys(spotify_ids) if sid]
    if not unique_ids:
        return

    query = select(Item.spotify_id).where(Item.spotify_id.in_(unique_ids))
    if excluded_item_ids:
        query = query.where(~Item.id.in_(excluded_item_ids))

    remaining_rows = (await db.execute(query)).all()
    remaining_ids = {row[0] for row in remaining_rows if row and row[0]}
    stale_ids = [sid for sid in unique_ids if sid not in remaining_ids]
    if stale_ids:
        await db.execute(delete(RawResponse).where(RawResponse.spotify_id.in_(stale_ids)))


def _item_to_response(
    item: Item,
    owner_url: str | None = None,
    raw_data: dict | None = None,
    snapshots: list[MetricsSnapshot] | None = None,
    item_user: User | None = None,
) -> ItemResponse:
    """Convert DB Item model to API response schema."""
    duration = None
    if item.duration_ms:
        mins, secs = divmod(item.duration_ms // 1000, 60)
        duration = f"{mins}:{secs:02d}"

    has_total_plays = item.item_type in {"track", "album", "playlist"}
    total_plays = item.playcount if has_total_plays else None
    artist_names = _extract_artist_names(raw_data)
    export_tracks = _build_album_export_tracks(raw_data) if item.item_type == "album" else None

    snapshots = snapshots or []
    latest_snapshot = snapshots[0] if len(snapshots) >= 1 else None
    previous_snapshot = snapshots[1] if len(snapshots) >= 2 else None

    delta_days = None
    if latest_snapshot and previous_snapshot:
        delta_days = max(
            0,
            int((latest_snapshot.captured_at - previous_snapshot.captured_at).total_seconds() // 86400),
        )

    return ItemResponse(
        id=str(item.id),
        spotify_id=item.spotify_id,
        type=item.item_type,
        name=item.name,
        spotify_url=_spotify_url(item.item_type, item.spotify_id),
        image=item.image,
        owner_name=item.owner_name,
        owner_image=item.owner_image,
        owner_url=owner_url,
        artist_names=artist_names or None,
        export_tracks=export_tracks or None,
        followers=item.followers,
        monthly_listeners=item.monthly_listeners,
        monthly_plays=item.monthly_listeners,
        playcount=item.playcount,
        total_plays=total_plays,
        track_count=item.track_count,
        album_count=item.album_count,
        duration=duration,
        release_date=item.release_date,
        saves=item.followers,
        followers_delta=_compute_delta(
            latest_snapshot.followers if latest_snapshot else item.followers,
            previous_snapshot.followers if previous_snapshot else None,
        ),
        monthly_listeners_delta=_compute_delta(
            latest_snapshot.monthly_listeners if latest_snapshot else item.monthly_listeners,
            previous_snapshot.monthly_listeners if previous_snapshot else None,
        ),
        playcount_delta=_compute_delta(
            latest_snapshot.playcount if latest_snapshot else item.playcount,
            previous_snapshot.playcount if previous_snapshot else None,
        ),
        track_count_delta=_compute_delta(
            latest_snapshot.track_count if latest_snapshot else item.track_count,
            previous_snapshot.track_count if previous_snapshot else None,
        ),
        delta_days=delta_days,
        status=item.status,
        error_code=item.error_code,
        error_message=item.error_message,
        group=item.group,
        user_id=str(item.user_id) if item.user_id else None,
        user_name=(item_user.display_name or item_user.username) if item_user else None,
        user_avatar=item_user.avatar if item_user else None,
        added_date=_format_added_date(item.created_at),
        last_checked=item.last_checked,
        created_at=item.created_at,
    )


def _safe_export_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def _format_export_metric(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return _safe_export_text(value)


def _format_export_updated(value: datetime | None) -> str:
    if value is None:
        return ""
    return value.strftime("%H:%M %d/%m/%Y")


def _build_export_file_name(prefix: str, extension: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    return f"{prefix}-{timestamp}.{extension}"


def _rows_to_tab_text(headers: list[str], rows: list[list[str]]) -> str:
    lines: list[str] = []
    if headers:
        lines.append("\t".join(_safe_export_text(cell) for cell in headers))
    for row in rows:
        lines.append("\t".join(_safe_export_text(cell) for cell in row))
    return "\r\n".join(lines)


def _extract_export_track_url(track: dict, fallback_type: str = "track") -> str:
    if not isinstance(track, dict):
        return ""
    spotify_url = track.get("spotify_url")
    if isinstance(spotify_url, str) and spotify_url.strip():
        return spotify_url.strip()
    spotify_id = track.get("spotify_id") or track.get("id")
    if isinstance(spotify_id, str) and spotify_id.strip():
        return _spotify_url(fallback_type, spotify_id.strip())
    return ""


def _extract_export_track_artists(track: dict, fallback: str = "-") -> str:
    if not isinstance(track, dict):
        return fallback

    artist_names = _extract_artist_names(track)
    if artist_names:
        return ", ".join(artist_names)

    owner_name = track.get("owner_name")
    if isinstance(owner_name, str) and owner_name.strip():
        return owner_name.strip()

    return fallback


def _build_export_track_title(artist_label: str, track_name: str) -> str:
    clean_track_name = _safe_export_text(track_name) or "-"
    clean_artist_label = _safe_export_text(artist_label) or ""
    if not clean_artist_label or clean_artist_label == "-":
        return clean_track_name
    prefix = f"{clean_artist_label} - "
    if clean_track_name.lower().startswith(prefix.lower()):
        return clean_track_name
    return f"{clean_artist_label} - {clean_track_name}"


def _merge_export_blocks(blocks: list[list[list[str]]]) -> list[list[str]]:
    normalized_blocks = [block for block in blocks if block]
    if not normalized_blocks:
        return []

    widths = [
        max((len(row) for row in block), default=0)
        for block in normalized_blocks
    ]
    height = max((len(block) for block in normalized_blocks), default=0)
    merged_rows: list[list[str]] = []

    for row_index in range(height):
        merged_row: list[str] = []
        for block_index, block in enumerate(normalized_blocks):
            width = widths[block_index]
            raw_row = block[row_index] if row_index < len(block) else []
            padded_row = [
                _safe_export_text(raw_row[column_index]) if column_index < len(raw_row) else ""
                for column_index in range(width)
            ]
            merged_row.extend(padded_row)
            if block_index < len(normalized_blocks) - 1:
                merged_row.append("")
        merged_rows.append(merged_row)

    return merged_rows


def _extract_playlist_owner(item: Item, raw_data: dict | None) -> str:
    if isinstance(raw_data, dict):
        owner_name = raw_data.get("owner_name")
        if isinstance(owner_name, str) and owner_name.strip():
            return owner_name.strip()
    if isinstance(item.owner_name, str) and item.owner_name.strip():
        return item.owner_name.strip()
    return ""


def _build_listview_export_rows(
    items: list[Item],
    raw_map: dict[str, dict],
    user_map: dict[str, User],
) -> tuple[list[str], list[list[str]], str]:
    headers = [
        "Type",
        "Name",
        "Spotify URL",
        "Group",
        "User",
        "Playlist Owner",
        "Playlist (Save)",
        "Playlist (Count)",
        "Album (Track Count)",
        "Artist (Followers)",
        "Artist (Listeners)",
        "Tracks (Views)",
        "Updated",
    ]
    rows: list[list[str]] = []
    for item in items:
        raw_data = raw_map.get(item.spotify_id)
        item_user = user_map.get(str(item.user_id)) if item.user_id else None
        user_name = ""
        if item_user:
            user_name = item_user.display_name or item_user.username or ""

        rows.append(
            [
                _safe_export_text(item.item_type),
                _safe_export_text(item.name),
                _spotify_url(item.item_type, item.spotify_id),
                _safe_export_text(item.group),
                _safe_export_text(user_name),
                _safe_export_text(_extract_playlist_owner(item, raw_data)),
                _format_export_metric(item.followers if item.item_type == "playlist" else None),
                _format_export_metric(item.track_count if item.item_type == "playlist" else None),
                _format_export_metric(item.track_count if item.item_type == "album" else None),
                _format_export_metric(item.followers if item.item_type in {"artist", "track", "album"} else None),
                _format_export_metric(item.monthly_listeners if item.item_type in {"artist", "track", "album"} else None),
                _format_export_metric(item.playcount if item.item_type == "track" else None),
                _format_export_updated(item.last_checked or item.created_at),
            ]
        )
    return headers, rows, "spoticheck-listview"


def _build_playlist_type3_rows(
    items: list[Item],
    raw_map: dict[str, dict],
) -> tuple[list[str], list[list[str]], str]:
    headers: list[str] = []
    blocks: list[list[list[str]]] = []
    for item in items:
        if item.item_type != "playlist":
            continue
        block_rows: list[list[str]] = [
            [_safe_export_text(item.name) or f"Playlist {item.spotify_id}", "", ""],
            ["Artist - Track", "Track Link", "PlayCount"],
        ]
        raw_data = raw_map.get(item.spotify_id)
        tracks = raw_data.get("tracks") if isinstance(raw_data, dict) else None
        if isinstance(tracks, list) and tracks:
            for track in tracks:
                if not isinstance(track, dict):
                    continue
                artist_label = _extract_export_track_artists(track)
                track_name = _safe_export_text(track.get("name")) or "-"
                title = _build_export_track_title(artist_label, track_name)
                block_rows.append(
                    [
                        title,
                        _extract_export_track_url(track),
                        _format_export_metric(track.get("playcount_estimate")),
                    ]
                )
        else:
            fallback_artist = _safe_export_text(item.owner_name) or "-"
            fallback_title = _build_export_track_title(fallback_artist, _safe_export_text(item.name) or "-")
            block_rows.append(
                [
                    fallback_title,
                    _spotify_url("playlist", item.spotify_id),
                    _format_export_metric(item.playcount),
                ]
            )
        blocks.append(block_rows)
    return headers, _merge_export_blocks(blocks), "spoticheck-playlist-type3"


def _build_album_type0_rows(
    items: list[Item],
    raw_map: dict[str, dict],
) -> tuple[list[str], list[list[str]], str]:
    headers: list[str] = []
    blocks: list[list[list[str]]] = []
    for item in items:
        if item.item_type != "album":
            continue
        album_name = _safe_export_text(item.name) or f"Album {item.spotify_id}"
        block_rows: list[list[str]] = [
            [album_name, "", "", ""],
            ["Track No", "Track Name", "Track Link", "PlayCount"],
        ]
        raw_data = raw_map.get(item.spotify_id)
        tracks = raw_data.get("tracks") if isinstance(raw_data, dict) else None
        if isinstance(tracks, list) and tracks:
            index = 1
            for track in tracks:
                if not isinstance(track, dict):
                    continue
                track_name = _safe_export_text(track.get("name")) or "-"
                artist_label = _extract_export_track_artists(track)
                block_rows.append(
                    [
                        str(index),
                        _build_export_track_title(artist_label, track_name),
                        _extract_export_track_url(track),
                        _format_export_metric(track.get("playcount_estimate")),
                    ]
                )
                index += 1
        else:
            fallback_artist = _safe_export_text(item.owner_name) or _safe_export_text(
                ", ".join(_extract_artist_names(raw_data))
            ) or "-"
            block_rows.append(
                [
                    "1",
                    _build_export_track_title(fallback_artist, _safe_export_text(item.name) or "-"),
                    _spotify_url("album", item.spotify_id),
                    _format_export_metric(item.playcount),
                ]
            )
        blocks.append(block_rows)
    return headers, _merge_export_blocks(blocks), "spoticheck-album-type0"


def _build_track_offline_rows(
    items: list[Item],
    raw_map: dict[str, dict],
) -> tuple[list[str], list[list[str]], str]:
    headers = ["Artist - Track", "Track Link", "PlayCount", "Listener/Month"]
    rows: list[list[str]] = []
    for item in items:
        if item.item_type != "track":
            continue
        raw_data = raw_map.get(item.spotify_id)
        artist_names = _extract_artist_names(raw_data)
        artist_label = ", ".join(artist_names) if artist_names else (_safe_export_text(item.owner_name) or "-")
        track_name = _safe_export_text(item.name) or "-"
        rows.append(
            [
                _build_export_track_title(artist_label, track_name),
                _spotify_url("track", item.spotify_id),
                _format_export_metric(item.playcount),
                _format_export_metric(item.monthly_listeners),
            ]
        )
    return headers, rows, "spoticheck-track-offline"


def _build_export_rows(
    action: str,
    items: list[Item],
    raw_map: dict[str, dict],
    user_map: dict[str, User],
) -> tuple[list[str], list[list[str]], str]:
    if action == "listview-excel":
        return _build_listview_export_rows(items, raw_map, user_map)
    if action == "playlist-type3":
        return _build_playlist_type3_rows(items, raw_map)
    if action == "album-type0":
        return _build_album_type0_rows(items, raw_map)
    if action == "track-offline":
        return _build_track_offline_rows(items, raw_map)
    raise HTTPException(status_code=400, detail="Unsupported export action")


async def _hydrate_raw_for_export(
    action: str,
    items: list[Item],
    raw_map: dict[str, dict],
    deep_fetch: bool,
) -> dict[str, dict]:
    if not deep_fetch:
        return raw_map
    if action not in {"playlist-type3", "album-type0"}:
        return raw_map

    target_items: list[Item] = []
    for item in items:
        has_tracks = isinstance(raw_map.get(item.spotify_id), dict) and isinstance(
            raw_map.get(item.spotify_id).get("tracks"),
            list,
        )
        if has_tracks:
            continue
        if action == "playlist-type3" and item.item_type == "playlist":
            target_items.append(item)
        if action == "album-type0" and item.item_type == "album":
            target_items.append(item)

    if not target_items:
        return raw_map

    semaphore = asyncio.Semaphore(4)

    async def fetch_one(target: Item) -> tuple[str, dict | None]:
        async with semaphore:
            if action == "playlist-type3":
                data = await spotify_client.fetch_playlist(target.spotify_id)
            else:
                data = await spotify_client.fetch_album(target.spotify_id)
            if isinstance(data, dict) and not data.get("error"):
                return target.spotify_id, data
            return target.spotify_id, None

    results = await asyncio.gather(
        *(fetch_one(target) for target in target_items),
        return_exceptions=True,
    )

    for result in results:
        if isinstance(result, Exception):
            continue
        spotify_id, payload = result
        if isinstance(payload, dict):
            raw_map[spotify_id] = payload

    return raw_map


def _build_xlsx_stream(headers: list[str], rows: list[list[str]]) -> io.BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required for XLSX export",
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"

    if headers:
        worksheet.append([_safe_export_text(cell) for cell in headers])
    for row in rows:
        worksheet.append([_safe_export_text(cell) for cell in row])

    for col_idx, column in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 64)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/items", response_model=ItemListResponse)
async def list_items(
    type: str | None = Query(None, description="Filter by item type"),
    group: str | None = Query(None, description="Filter by group"),
    status: str | None = Query(None, description="Filter by status"),
    user_id: str | None = Query(None, description="Filter by user (admin only)"),
    limit: int = Query(100, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all items with optional filters."""
    query = select(Item).order_by(Item.updated_at.desc())

    if current_user.role != "admin":
        query = query.where(Item.user_id == current_user.id)
    elif user_id:
        query = query.where(Item.user_id == user_id)

    if type:
        query = query.where(Item.item_type == type)
    if group:
        query = query.where(Item.group == group)
    if status:
        query = query.where(Item.status == status)

    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    query = query.limit(limit).offset(offset)
    result = await db.execute(query)
    items = result.scalars().all()
    raw_map = await _load_latest_raw_data(db, items)
    snapshot_map = await _load_recent_snapshots(db, items)
    user_map = await _load_item_users(db, items)

    return ItemListResponse(
        items=[
            _item_to_response(
                item,
                _extract_owner_url(item, raw_map.get(item.spotify_id)),
                raw_map.get(item.spotify_id),
                snapshot_map.get(item.id),
                user_map.get(str(item.user_id)) if item.user_id else None,
            )
            for item in items
        ],
        total=total,
    )


@router.get("/items/{item_type}/{spotify_id}", response_model=ItemResponse)
async def get_item(
    item_type: str,
    spotify_id: str,
    user_id: str | None = Query(None, description="Target user (admin only)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get a specific item by type and Spotify ID."""
    query = select(Item).where(
        Item.spotify_id == spotify_id,
        Item.item_type == item_type,
    )
    if current_user.role != "admin":
        query = query.where(Item.user_id == current_user.id)
    elif user_id:
        query = query.where(Item.user_id == user_id)

    result = await db.execute(query.order_by(Item.updated_at.desc()))
    rows = result.scalars().all()
    if current_user.role == "admin" and not user_id and len(rows) > 1:
        raise HTTPException(
            status_code=409,
            detail="Multiple users track this link. Specify user_id.",
        )
    item = rows[0] if rows else None
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    if current_user.role != "admin" and item.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to view this item")

    raw_map = await _load_latest_raw_data(db, [item])
    snapshot_map = await _load_recent_snapshots(db, [item])
    user_map = await _load_item_users(db, [item])
    return _item_to_response(
        item,
        _extract_owner_url(item, raw_map.get(item.spotify_id)),
        raw_map.get(item.spotify_id),
        snapshot_map.get(item.id),
        user_map.get(str(item.user_id)) if item.user_id else None,
    )


@router.patch("/items/group")
async def rename_group(
    old_group: str = Query(..., description="Current group name"),
    new_group: str | None = Query(None, description="New group name (empty = clear group)"),
    user_id: str | None = Query(None, description="Target user (admin only)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Rename or clear a group for one user and persist item.group in DB."""
    old_clean = old_group.strip()
    new_clean = (new_group or "").strip()
    if not old_clean:
        raise HTTPException(status_code=400, detail="old_group is required")
    next_group = new_clean or None
    old_norm = _normalize_group_name(old_clean)

    target_user_id = current_user.id
    if current_user.role == "admin" and user_id:
        try:
            target_user_id = uuid.UUID(str(user_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid user_id") from exc

    query = select(Item).where(Item.group.is_not(None))
    if current_user.role == "admin":
        if str(target_user_id) == str(current_user.id):
            # Legacy admin rows may still have null user_id.
            query = query.where(or_(Item.user_id == target_user_id, Item.user_id.is_(None)))
        else:
            query = query.where(Item.user_id == target_user_id)
    else:
        query = query.where(Item.user_id == target_user_id)

    items = (await db.execute(query)).scalars().all()
    updated = 0
    for item in items:
        if _normalize_group_name(item.group) != old_norm:
            continue
        item.group = next_group
        updated += 1

    await db.commit()

    return {
        "ok": True,
        "updated": int(updated),
        "old_group": old_clean,
        "new_group": next_group,
    }


@router.delete("/items/{item_type}/{spotify_id}")
async def delete_item(
    item_type: str,
    spotify_id: str,
    user_id: str | None = Query(None, description="Target user (admin only)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a single item by type + Spotify ID."""
    query = select(Item).where(
        Item.spotify_id == spotify_id,
        Item.item_type == item_type,
    )
    if current_user.role != "admin":
        query = query.where(Item.user_id == current_user.id)
    elif user_id:
        query = query.where(Item.user_id == user_id)

    result = await db.execute(query.order_by(Item.updated_at.desc()))
    rows = result.scalars().all()
    if current_user.role == "admin" and not user_id and len(rows) > 1:
        raise HTTPException(
            status_code=409,
            detail="Multiple users track this link. Specify user_id.",
        )
    item = rows[0] if rows else None
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    if current_user.role != "admin" and item.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this item")

    await db.execute(delete(MetricsSnapshot).where(MetricsSnapshot.item_id == item.id))
    await db.execute(delete(CrawlJob).where(CrawlJob.item_id == item.id))
    await db.delete(item)
    await _delete_raw_if_unreferenced(db, [item.spotify_id], excluded_item_ids=[item.id])
    await db.commit()
    return {"ok": True, "deleted": 1}


@router.delete("/items-by-id/{item_id}")
async def delete_item_by_id(
    item_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a single item by row id (safe when duplicate links exist)."""
    try:
        item_uuid = uuid.UUID(str(item_id))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid item_id") from exc

    result = await db.execute(select(Item).where(Item.id == item_uuid))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    if current_user.role != "admin" and item.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this item")

    await db.execute(delete(MetricsSnapshot).where(MetricsSnapshot.item_id == item.id))
    await db.execute(delete(CrawlJob).where(CrawlJob.item_id == item.id))
    await db.delete(item)
    await _delete_raw_if_unreferenced(db, [item.spotify_id], excluded_item_ids=[item.id])
    await db.commit()
    return {"ok": True, "deleted": 1}


@router.delete("/items")
async def clear_items(
    group: str | None = Query(None, description="Optional group to clear"),
    user_id: str | None = Query(None, description="Filter by user (admin only)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Clear all items (or items in one group)."""
    selected_items = select(Item.id, Item.spotify_id)

    if current_user.role != "admin":
        selected_items = selected_items.where(Item.user_id == current_user.id)
    elif user_id:
        selected_items = selected_items.where(Item.user_id == user_id)

    if group:
        selected_items = selected_items.where(Item.group == group)

    rows = (await db.execute(selected_items)).all()
    if not rows:
        return {"ok": True, "deleted": 0, "group": group}

    item_ids = [row[0] for row in rows]
    spotify_ids = [row[1] for row in rows]

    await db.execute(delete(MetricsSnapshot).where(MetricsSnapshot.item_id.in_(item_ids)))
    await db.execute(delete(CrawlJob).where(CrawlJob.item_id.in_(item_ids)))
    result = await db.execute(delete(Item).where(Item.id.in_(item_ids)))
    await _delete_raw_if_unreferenced(db, spotify_ids, excluded_item_ids=item_ids)
    await db.commit()
    return {"ok": True, "deleted": result.rowcount or 0, "group": group}


@router.post("/items/export")
async def export_items(
    payload: ItemExportRequest = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    action = _safe_export_text(payload.action).lower()
    output_format = _safe_export_text(payload.format).lower() or "json"
    if action not in EXPORT_ACTIONS:
        raise HTTPException(status_code=400, detail="Unsupported export action")
    if output_format not in EXPORT_FORMATS:
        raise HTTPException(status_code=400, detail="Unsupported export format")

    ordered_item_ids: list[str] = []
    ordered_item_uuids: list[uuid.UUID] = []
    seen_ids: set[str] = set()
    for raw_item_id in payload.item_ids:
        item_id = _safe_export_text(raw_item_id)
        if not item_id or item_id in seen_ids:
            continue
        try:
            parsed = uuid.UUID(item_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid item id: {item_id}") from exc
        seen_ids.add(item_id)
        ordered_item_ids.append(item_id)
        ordered_item_uuids.append(parsed)

    if not ordered_item_uuids:
        raise HTTPException(status_code=400, detail="No item_ids provided")

    query = select(Item).where(Item.id.in_(ordered_item_uuids))
    if current_user.role != "admin":
        query = query.where(Item.user_id == current_user.id)
    result = await db.execute(query)
    rows = result.scalars().all()
    if not rows:
        raise HTTPException(status_code=404, detail="No items found for export")

    item_map = {str(item.id): item for item in rows}
    items: list[Item] = [item_map[item_id] for item_id in ordered_item_ids if item_id in item_map]
    if not items:
        raise HTTPException(status_code=404, detail="No exportable items found")
    if len(items) < len(ordered_item_ids):
        raise HTTPException(status_code=404, detail="Some items were not found or not authorized")

    raw_map = await _load_latest_raw_data(db, items)
    raw_map = await _hydrate_raw_for_export(action, items, raw_map, bool(payload.deep_fetch))
    user_map = await _load_item_users(db, items)
    headers, export_rows, file_prefix = _build_export_rows(action, items, raw_map, user_map)

    if output_format == "json":
        return {
            "ok": True,
            "action": action,
            "headers": headers,
            "rows": export_rows,
            "count": len(export_rows),
            "file_prefix": file_prefix,
        }

    if output_format == "txt":
        text_content = _rows_to_tab_text(headers, export_rows)
        file_name = _build_export_file_name(file_prefix, "txt")
        response_headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
        return StreamingResponse(
            io.BytesIO(text_content.encode("utf-8-sig")),
            media_type="text/plain; charset=utf-8",
            headers=response_headers,
        )

    if output_format == "xlsx":
        stream = _build_xlsx_stream(headers, export_rows)
        file_name = _build_export_file_name(file_prefix, "xlsx")
        response_headers = {"Content-Disposition": f'attachment; filename="{file_name}"'}
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=response_headers,
        )

    raise HTTPException(status_code=400, detail="Unsupported export format")
